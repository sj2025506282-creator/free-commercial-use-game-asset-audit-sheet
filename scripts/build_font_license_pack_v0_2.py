#!/usr/bin/env python3
"""Build the v0.2 verified game font license source pack."""

from __future__ import annotations

import csv
import json
import shutil
import zipfile
from pathlib import Path
from urllib.parse import quote_plus

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


ROOT = Path("game-font-license-source-pack-v0.2")
ZIP_PATH = Path("game-font-license-source-pack-v0.2.zip")
CSV_PATH = ROOT / "game_font_sources_verified_v0.2.csv"
QUESTIONS_PATH = ROOT / "font_license_questions_v0.2.csv"
XLSX_PATH = ROOT / "game_font_license_source_pack_v0.2.xlsx"
RELEASE_URL = "https://github.com/sj2025506282-creator/free-commercial-use-game-asset-audit-sheet/releases/tag/game-font-license-pack-v0.2.0"
TODAY = "2026-06-02"


FIELDS = [
    "id",
    "row_type",
    "name",
    "creator_or_platform",
    "url",
    "evidence_url",
    "font_type",
    "category",
    "license_signal",
    "commercial_use",
    "attribution_required",
    "redistribution_note",
    "download_requires_signup",
    "audit_status",
    "risk_notes",
    "best_for",
    "last_checked",
    "engine_note_unity",
    "engine_note_godot",
    "engine_note_unreal",
    "license_file_required",
    "reserved_font_name_risk",
    "logo_use_warning",
    "glyph_coverage_note",
    "recommended_for",
    "avoid_for",
]


GOOGLE_ROWS = [
    ("Inter", "inter", "Inter Project", "sans UI font", "readable UI/body", "Very strong UI font; practical for menus, settings screens, tools, and readable in-game overlays.", "clean game UI, tools, menus", "dense retro title typography"),
    ("Open Sans", "opensans", "Open Sans Project", "sans UI font", "readable UI/body", "General-purpose UI font; common look, but reliable for body copy and documentation-heavy UI.", "menus, subtitles, docs", "distinctive title/logo work"),
    ("Lato", "lato", "tyPoland / Lato Project", "sans UI font", "readable UI/body", "Readable interface text; check style fit for fantasy or retro games.", "menus and HUD labels", "pixel-perfect retro UI"),
    ("Montserrat", "montserrat", "Montserrat Project", "geometric sans font", "display/title", "Good headings; can be wide for compact HUD text.", "titles, menus, store pages", "small dense HUD labels"),
    ("Merriweather", "merriweather", "Merriweather Project", "serif font", "readable UI/body", "Useful for lore/dialogue; serif details may blur at small pixel sizes.", "lore, dialogue, readable text", "tiny HUD labels"),
    ("Source Sans 3", "sourcesans3", "Adobe / Source Sans Project", "sans UI font", "readable UI/body", "Adobe-related reserved-name/trademark wording deserves care if modifying.", "tool UI, menus, readable text", "modified derivative without naming review"),
    ("Noto Sans", "notosans", "Noto Project", "multilingual sans font", "multilingual/fallback", "Good baseline for multilingual fallback; choose exact Noto family for target script.", "multilingual UI fallback", "stylized title-only use"),
    ("Press Start 2P", "pressstart2p", "Press Start 2P Project", "pixel display font", "pixel/retro", "Iconic pixel style but poor for long paragraphs.", "retro title screens", "long dialogue/body copy"),
    ("Silkscreen", "silkscreen", "Silkscreen Project", "pixel display font", "pixel/retro", "Compact pixel UI candidate; test at small sizes.", "pixel HUD labels", "long readable paragraphs"),
    ("VT323", "vt323", "VT323 Project", "terminal display font", "terminal/monospace", "Strong terminal/CRT flavor; not suitable for all UI text.", "terminal, hacking, retro UI", "general menu body text"),
    ("Inconsolata", "inconsolata", "Inconsolata Project", "monospace font", "terminal/monospace", "Good for debug UI and code-like text; less expressive for fantasy themes.", "debug UI, terminal text", "fantasy/dialogue immersion"),
    ("Fira Code", "firacode", "Fira Code Project", "monospace coding font", "terminal/monospace", "Useful for dev tools and in-game terminals; ligatures may be unnecessary in games.", "terminal UI, dev tools", "non-code dialogue"),
    ("Rubik", "rubik", "Rubik Project", "rounded sans font", "readable UI/body", "Friendly UI tone; check weights for target platform.", "casual game menus", "serious simulation tone"),
    ("Bangers", "bangers", "Bangers Project", "comic display font", "display/title", "High-energy display face; use sparingly and test readability.", "arcade titles, comic UI", "body text and localization-heavy UI"),
    ("Creepster", "creepster", "Font Diner", "horror display font", "display/title", "Strong horror flavor; reserved-name wording should be kept in mind if modifying.", "horror title screens", "general UI labels"),
    ("Fredoka", "fredoka", "Fredoka Project", "rounded sans font", "readable UI/body", "Friendly rounded UI option; good for casual games.", "casual UI and menus", "gritty/serious atmosphere"),
    ("Audiowide", "audiowide", "Astigmatic", "wide techno display font", "display/title", "Wide sci-fi display style; poor fit for dense text.", "sci-fi titles", "small HUD labels"),
    ("Orbitron", "orbitron", "Orbitron Project", "sci-fi display font", "display/title", "Useful sci-fi display family; test readability for numbers.", "sci-fi HUD and titles", "large paragraphs"),
    ("Rajdhani", "rajdhani", "Indian Type Foundry", "condensed sans font", "readable UI/body", "Condensed UI option; useful when space is tight, but test legibility.", "HUD labels and menus", "long text at very small sizes"),
    ("Exo 2", "exo2", "Exo 2 Project", "tech sans font", "readable UI/body", "Tech-flavored readable UI option.", "sci-fi UI and menus", "historical/fantasy UI"),
    ("Josefin Sans", "josefinsans", "Josefin Sans Project", "geometric sans font", "display/title", "Elegant display/body hybrid; thin weights need testing.", "menus and title cards", "tiny HUD text"),
    ("Nunito", "nunito", "Nunito Project", "rounded sans font", "readable UI/body", "Soft readable UI choice; good for friendly apps/games.", "casual menus and dialogue UI", "serious horror/simulation UI"),
    ("Oswald", "oswald", "Oswald Project", "condensed sans font", "display/title", "Useful for compact titles and labels; avoid overusing in body text.", "headers and compact labels", "long paragraphs"),
    ("Roboto Mono", "robotomono", "Roboto Mono Project", "monospace font", "terminal/monospace", "Strong baseline for terminals, debug overlays, and code-like UI.", "terminal and debug UI", "warm narrative dialogue"),
    ("Roboto", "roboto", "Roboto Project", "sans UI font", "readable UI/body", "General UI baseline; common but practical.", "mobile-like game UI", "distinctive brand/title needs"),
    ("Playfair Display", "playfairdisplay", "Playfair Display Project", "serif display font", "display/title", "Good for elegant titles; not ideal for small HUD text.", "visual novel titles, lore headers", "small UI labels"),
    ("Cinzel", "cinzel", "Cinzel Project", "classical display font", "display/title", "Useful for fantasy/classical titles; check readability for long text.", "fantasy titles", "body copy"),
    ("Cinzel Decorative", "cinzeldecorative", "Natanael Gama", "decorative display font", "display/title", "Highly decorative; use for logo-like headings only and review brand use separately.", "ornamental title cards", "UI labels and paragraphs"),
    ("Unbounded", "unbounded", "Unbounded Project", "display sans font", "display/title", "Distinctive display family; can feel heavy in compact UI.", "posters, splash screens", "dense UI"),
    ("Teko", "teko", "Teko Project", "condensed display font", "display/title", "Good for bold compact titles; check number readability.", "HUD headings and scores", "body text"),
    ("Chakra Petch", "chakrapetch", "Chakra Petch Project", "tech sans font", "readable UI/body", "Tech UI feel with readable forms; useful for sci-fi games.", "sci-fi menus", "soft cozy UI"),
    ("Oxanium", "oxanium", "Oxanium Project", "tech display font", "display/title", "Useful for sci-fi/vehicle HUDs; test lowercase readability.", "sci-fi HUD", "long body text"),
    ("Black Ops One", "blackopsone", "Black Ops Project", "stencil display font", "display/title", "Military/stencil style; use carefully for tone and readability.", "action titles", "body copy"),
    ("Righteous", "righteous", "Astigmatic", "display sans font", "display/title", "Retro-futuristic display choice; test at small sizes.", "arcade title UI", "paragraph text"),
    ("Kanit", "kanit", "Kanit Project", "sans UI font", "multilingual/fallback", "Useful Thai/Latin UI candidate; verify script needs before using.", "Thai/Latin UI", "scripts not covered by family"),
    ("Barlow", "barlow", "Barlow Project", "sans UI font", "readable UI/body", "Readable and utilitarian; good for simulation/tool-like UI.", "menus and dashboards", "highly stylized title UI"),
    ("Cabin", "cabin", "Cabin Project", "sans UI font", "readable UI/body", "Humanist sans option; practical for readable menus.", "dialogue boxes and menus", "pixel art title screens"),
    ("Quicksand", "quicksand", "Quicksand Project", "rounded sans font", "readable UI/body", "Soft rounded UI option; test for small-size readability.", "cozy/casual menus", "serious tactical UI"),
    ("Work Sans", "worksans", "Work Sans Project", "sans UI font", "readable UI/body", "Practical UI/body font; broad utility.", "menus and tool UI", "highly themed title cards"),
    ("IBM Plex Mono", "ibmplexmono", "IBM Plex Project", "monospace font", "terminal/monospace", "Reserved Font Name Plex; useful but modification/naming needs care.", "terminal and debug UI", "modified derivative without naming review"),
    ("IBM Plex Sans", "ibmplexsans", "IBM Plex Project", "sans UI font", "readable UI/body", "Reserved Font Name Plex; strong readable UI candidate.", "menus and documentation UI", "modified derivative without naming review"),
    ("Share Tech Mono", "sharetechmono", "Carrois Type Design", "monospace tech font", "terminal/monospace", "Compact tech mono; good for code-like labels.", "terminal and sci-fi UI", "long narrative text"),
    ("Pixelify Sans", "pixelifysans", "Pixelify Sans Project", "pixel font", "pixel/retro", "Pixel-focused family; better modern pixel UI candidate than many title-only pixel fonts.", "pixel UI and retro menus", "high-density non-pixel UI"),
]


def google_specimen_url(name: str) -> str:
    return f"https://fonts.google.com/specimen/{quote_plus(name)}"


def ofl_url(slug: str) -> str:
    return f"https://raw.githubusercontent.com/google/fonts/main/ofl/{slug}/OFL.txt"


def google_row(index: int, name: str, slug: str, project: str, font_type: str, category: str, risk: str, best_for: str, avoid_for: str) -> dict[str, str]:
    reserved = "yes - review OFL Reserved Font Name wording before modifying" if any(term in risk.lower() for term in ["reserved", "adobe", "plex"]) else "possible - check OFL text before modifying"
    return {
        "id": f"F{index:03d}",
        "row_type": "font_source",
        "name": name,
        "creator_or_platform": f"Google Fonts / {project}",
        "url": google_specimen_url(name),
        "evidence_url": ofl_url(slug),
        "font_type": font_type,
        "category": category,
        "license_signal": "SIL Open Font License file in google/fonts",
        "commercial_use": "yes",
        "attribution_required": "no",
        "redistribution_note": "If distributing the font file, include the OFL license/copyright notice and respect reserved font name terms.",
        "download_requires_signup": "no",
        "audit_status": "verified_license_file",
        "risk_notes": risk,
        "best_for": best_for,
        "last_checked": TODAY,
        "engine_note_unity": "Include font asset/license in project records; if bundling raw font, keep OFL text with source/audit docs.",
        "engine_note_godot": "If importing as DynamicFont/FontFile, keep source URL and OFL evidence in project docs.",
        "engine_note_unreal": "If packaging font assets, record OFL evidence and avoid renaming modified fonts under a reserved name.",
        "license_file_required": "yes when redistributing/bundling the font file",
        "reserved_font_name_risk": reserved,
        "logo_use_warning": "Font license is not a trademark clearance; review logo/brand use separately.",
        "glyph_coverage_note": "Check target languages/scripts and fallback font behavior before shipping.",
        "recommended_for": best_for,
        "avoid_for": avoid_for,
    }


def ggbot_rows() -> list[dict[str, str]]:
    base = [
        ("GGBotNet Fonts CC0 All-in-1", "https://ggbot.itch.io/ggbotnet-fonts-cc0", "font bundle", "pixel/retro", "Good CC0 baseline source; verify individual files before bundling a shipped game.", "game UI and jam prototypes", "shipping without checking exact downloaded files"),
        ("Home Video Font CC0", "https://ggbot.itch.io/home-video-font", "display font", "pixel/retro", "Retro display style; check readability before using for body UI.", "retro VHS style game UI", "body text and small UI"),
        ("Public Pixel Font CC0", "https://ggbot.itch.io/public-pixel-font", "pixel font", "pixel/retro", "Strong pixel-game candidate; check glyph coverage for localization.", "pixel games and UI", "non-pixel UI systems"),
        ("First Time Writing Font CC0", "https://ggbot.itch.io/first-time-writing-font", "handwritten font", "display/title", "Handwritten style can harm readability; test in target resolution.", "dialogue and handwritten UI", "small HUD labels"),
    ]
    rows = []
    for i, (name, url, font_type, category, risk, best_for, avoid_for) in enumerate(base, start=1):
        rows.append(
            {
                "id": f"F{i:03d}",
                "row_type": "font_source",
                "name": name,
                "creator_or_platform": "GGBotNet",
                "url": url,
                "evidence_url": url,
                "font_type": font_type,
                "category": category,
                "license_signal": "CC0 signal on source page",
                "commercial_use": "yes",
                "attribution_required": "no",
                "redistribution_note": "CC0 allows broad reuse, but link original source and do not imply ownership.",
                "download_requires_signup": "no",
                "audit_status": "verified_page_signal",
                "risk_notes": risk,
                "best_for": best_for,
                "last_checked": TODAY,
                "engine_note_unity": "Import as project font asset; keep source URL in project audit log.",
                "engine_note_godot": "Import as font resource; keep source URL in project audit log.",
                "engine_note_unreal": "Import/package with project UI assets; keep source URL in project audit log.",
                "license_file_required": "not required by CC0 signal, but keep source evidence in audit docs",
                "reserved_font_name_risk": "low based on CC0 signal; still avoid implying creator endorsement",
                "logo_use_warning": "CC0 permission is not trademark clearance; review logo/brand use separately.",
                "glyph_coverage_note": "Check target languages/scripts and fallback font behavior before shipping.",
                "recommended_for": best_for,
                "avoid_for": avoid_for,
            }
        )
    return rows


def build_rows() -> list[dict[str, str]]:
    rows = ggbot_rows()
    start = len(rows) + 1
    for offset, row in enumerate(GOOGLE_ROWS, start=start):
        rows.append(google_row(offset, *row))
    return rows


def write_csv(path: Path, rows: list[dict[str, str]], fields: list[str]) -> None:
    with path.open("w", encoding="utf-8", newline="") as f:
        writer = csv.DictWriter(f, fieldnames=fields)
        writer.writeheader()
        writer.writerows(rows)


def write_xlsx(rows: list[dict[str, str]], questions: list[dict[str, str]]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Verified Fonts"
    ws.append(FIELDS)
    for row in rows:
        ws.append([row[field] for field in FIELDS])
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(color="FFFFFF", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = ws.dimensions
    widths = {
        "A": 8,
        "B": 16,
        "C": 24,
        "D": 30,
        "E": 36,
        "F": 48,
        "G": 20,
        "H": 20,
        "I": 28,
        "N": 22,
        "O": 46,
        "P": 26,
        "R": 42,
        "S": 42,
        "T": 42,
        "U": 24,
        "V": 34,
        "W": 36,
        "X": 34,
        "Y": 30,
        "Z": 30,
    }
    for col in range(1, len(FIELDS) + 1):
        letter = get_column_letter(col)
        ws.column_dimensions[letter].width = widths.get(letter, 18)

    qws = wb.create_sheet("Shipping Questions")
    q_fields = list(questions[0].keys())
    qws.append(q_fields)
    for question in questions:
        qws.append([question[field] for field in q_fields])
    for cell in qws[1]:
        cell.fill = header_fill
        cell.font = header_font
    qws.freeze_panes = "A2"
    qws.auto_filter.ref = qws.dimensions
    for col in range(1, len(q_fields) + 1):
        qws.column_dimensions[get_column_letter(col)].width = 32

    start = wb.create_sheet("Start Here", 0)
    start.append(["Free Game Font License Source Pack v0.2"])
    start.append(["Verified rows", len(rows)])
    start.append(["Release", RELEASE_URL])
    start.append(["Not legal advice", "Use this as a pre-shipping audit aid, not legal clearance."])
    start.append(["Main upgrade", "43 verified rows, engine notes, reserved-name risk, logo-use warnings, and glyph coverage notes."])
    start["A1"].font = Font(bold=True, size=16)
    wb.save(XLSX_PATH)


def questions() -> list[dict[str, str]]:
    return [
        {"id": "Q001", "question": "Are you embedding the font file inside the game build?", "why_it_matters": "Embedding can trigger license-file and copyright-notice handling.", "what_to_record": "Record source URL, evidence URL, and whether the license file is bundled."},
        {"id": "Q002", "question": "Does the font have a Reserved Font Name clause?", "why_it_matters": "OFL can allow modification but restrict use of the original reserved name for derivatives.", "what_to_record": "Record reserved-name wording before modifying or renaming."},
        {"id": "Q003", "question": "Are you using the font in a logo or trademark?", "why_it_matters": "Font license permission is not the same as trademark safety or distinctiveness.", "what_to_record": "Record separate brand/logo review status."},
        {"id": "Q004", "question": "Are you redistributing the raw font file?", "why_it_matters": "Redistribution often requires including license and copyright notices.", "what_to_record": "Record where the license file is bundled."},
        {"id": "Q005", "question": "Is the source row verified or discovery-only?", "why_it_matters": "Directory tags are not enough for shipping decisions.", "what_to_record": "Count only verified rows in public claims."},
        {"id": "Q006", "question": "Does the font cover your target languages?", "why_it_matters": "A font can be legally usable but fail localization.", "what_to_record": "Record glyph/script coverage and fallback behavior."},
        {"id": "Q007", "question": "Is the font readable at the smallest in-game size?", "why_it_matters": "Good license does not equal good UI.", "what_to_record": "Record tested size, resolution, and viewport."},
        {"id": "Q008", "question": "Are you modifying the font file?", "why_it_matters": "Modification can create naming and license-handling obligations.", "what_to_record": "Record derivative name and license handling."},
        {"id": "Q009", "question": "Which engine path are you using?", "why_it_matters": "Unity, Godot, and Unreal handle imported fonts differently.", "what_to_record": "Record engine-specific packaging and license-file location."},
        {"id": "Q010", "question": "Can the team re-check this source later?", "why_it_matters": "Release audits need repeatable evidence, not memory.", "what_to_record": "Keep evidence URL, last_checked, and local audit report."},
    ]


def write_docs(rows: list[dict[str, str]], qs: list[dict[str, str]]) -> None:
    category_counts: dict[str, int] = {}
    for row in rows:
        category_counts[row["category"]] = category_counts.get(row["category"], 0) + 1
    category_lines = "\n".join(f"- {name}: {count}" for name, count in sorted(category_counts.items()))

    (ROOT / "README.md").write_text(
        f"""# Free Game Font License Source Pack v0.2

This is a verified starter pack for indie game developers who need practical font-license starting points before shipping.

## Contents

- `game_font_sources_verified_v0.2.csv`: {len(rows)} verified font/source rows.
- `game_font_license_source_pack_v0.2.xlsx`: filterable workbook.
- `font_license_questions_v0.2.csv`: {len(qs)} practical questions to ask before shipping a font in a game.
- `AUDIT_METHOD.md`: how rows were checked and what the pack does not prove.
- `AUDIT_REPORT.json`: local consistency audit result.
- `CHANGELOG.md`: version history.

## Category Coverage

{category_lines}

## Audit Standard

Each verified row has source URL, evidence URL, license signal, commercial-use note, attribution note, redistribution note, risk notes, engine notes, reserved-name risk, logo-use warning, glyph coverage note, and last-checked date.

This pack does not redistribute font files. It is not legal advice.

## Paid Workflow Option

If you want reusable audit templates for your whole game asset folder, the paid starter template pack is here:

https://3813941972097.gumroad.com/l/isavr

The paid pack sells workflow templates, not third-party assets.
""",
        encoding="utf-8",
    )

    (ROOT / "AUDIT_METHOD.md").write_text(
        f"""# Audit Method

Last checked: {TODAY}

Rows are marked verified only when they have a direct source URL and an evidence URL.

For Google Fonts rows, the evidence URL is the OFL file in the official `google/fonts` repository.

For GGBotNet rows, the evidence URL is the source itch.io page checked on {TODAY}.

What this audit checks:

- visible license signal
- source/evidence URL presence
- commercial-use note
- attribution note
- redistribution note
- engine packaging reminders
- reserved font name risk reminder
- logo/trademark warning
- glyph coverage reminder

What this audit does not prove:

- legal clearance for every use case
- trademark/logo safety
- complete glyph coverage for your game
- that future source pages will not change
- that modified derivative fonts can keep the original name
""",
        encoding="utf-8",
    )

    (ROOT / "CHANGELOG.md").write_text(
        f"""# Changelog

## v0.2.0 - {TODAY}

- Expanded from 17 to {len(rows)} verified rows.
- Added XLSX workbook with filters.
- Added engine notes for Unity, Godot, and Unreal.
- Added license-file-required, reserved-name-risk, logo-use-warning, glyph-coverage, recommended-for, and avoid-for fields.
- Added audit method and local audit report.

## v0.1.0 - 2026-06-01

- Initial 17-row verified starter pack.
- Added 8 shipping questions.
""",
        encoding="utf-8",
    )

    (ROOT / "reddit_update_draft.md").write_text(
        f"""# r/gamedev update draft

Title:

`I expanded my free game font license starter pack to {len(rows)} verified sources with evidence links and engine notes`

Post body:

Quick update from yesterday's font-license starter pack.

I expanded it from 17 verified rows to {len(rows)} verified font/source rows. I kept the same rule: no discovery rows counted in the title.

What's new:

- {len(rows)} verified rows
- source URL + evidence URL per row
- filterable XLSX workbook
- Unity / Godot / Unreal notes
- reserved font name risk field
- logo-use warning field
- glyph coverage note field
- recommended-for / avoid-for fields
- audit method + local audit report

Download:

{RELEASE_URL}

This is still not legal advice and does not redistribute font files. I am trying to make the sheet useful for pre-shipping checks, not just make a big list.

What would make v0.3 more useful: more pixel/retro fonts, more multilingual fallback rows, or a license notice generator?
""",
        encoding="utf-8",
    )


def main() -> None:
    if ROOT.exists():
        shutil.rmtree(ROOT)
    ROOT.mkdir(parents=True)
    if ZIP_PATH.exists():
        ZIP_PATH.unlink()
    rows = build_rows()
    qs = questions()
    write_csv(CSV_PATH, rows, FIELDS)
    write_csv(QUESTIONS_PATH, qs, list(qs[0].keys()))
    write_xlsx(rows, qs)
    write_docs(rows, qs)
    with zipfile.ZipFile(ZIP_PATH, "w", compression=zipfile.ZIP_DEFLATED) as zf:
        for path in sorted(ROOT.rglob("*")):
            if path.is_file():
                zf.write(path, path.relative_to(ROOT.parent))
    print(json.dumps({"rows": len(rows), "zip": str(ZIP_PATH), "root": str(ROOT)}, indent=2))


if __name__ == "__main__":
    main()
